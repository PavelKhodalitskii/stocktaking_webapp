import pandas as pd
from openpyxl import Workbook
from items_management.models import InventoryItems
from reports.models import StocktalkingReport


def create_report(report_id = 1):
    try:
        report = StocktalkingReport.objects.get(id=report_id)
        items = report.items.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по инвентарю"

        ws.cell(row=1, column=15, value="Примечание")
        ws.cell(row=1, column=14, value="Материально ответственное лицо")
        ws.cell(row=1, column=10, value="Результаты инвентаризации")
        ws.cell(row=1, column=8, value="По данным бухгалтерского учета")
        ws.cell(row=1, column=4, value="Фактическое наличие")
        ws.cell(row=1, column=1, value="N")
        ws.cell(row=1, column=2, value="Наименование объекта нефинансового учета")
        ws.cell(row=1, column=3, value="Номер(код) объекта учета")
        ws.cell(row=2, column=4, value="Цена(Оценочная стоимость)")
        ws.cell(row=2, column=5, value="Количество")
        ws.cell(row=2, column=6, value="Сумма, руб")
        ws.cell(row=2, column=7, value="Статус объекта учета")
        ws.cell(row=2, column=8, value="Количество")
        ws.cell(row=2, column=9, value="Балансовая стоимость, руб")
        ws.cell(row=2, column=10, value="Отклонение")
        ws.cell(row=3, column=4, value="Цена(оценочная стоимость), руб")
        ws.cell(row=3, column=5, value="Количество")
        ws.cell(row=3, column=6, value="Сумма, руб")
        ws.cell(row=3, column=7, value="Статус объекта учета")
        ws.cell(row=3, column=8, value="Количество")
        ws.cell(row=3, column=9, value="Балансовая стоимость, руб")
        ws.cell(row=3, column=10, value="Недосдача")
        ws.cell(row=3, column=12, value="Излишки")
        ws.cell(row=4, column=10, value="Количество")
        ws.cell(row=4, column=11, value="Сумма, руб")
        ws.cell(row=4, column=12, value="Количество")
        ws.cell(row=4, column=13, value="Сумма, руб")


        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=4, end_column=2)
        ws.merge_cells(start_row=1, start_column=3, end_row=4, end_column=3)
        ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=7)
        ws.merge_cells(start_row=3, start_column=4, end_row=4, end_column=4)
        ws.merge_cells(start_row=3, start_column=5, end_row=4, end_column=5)
        ws.merge_cells(start_row=3, start_column=6, end_row=4, end_column=6)
        ws.merge_cells(start_row=3, start_column=7, end_row=4, end_column=7)
        ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
        ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=8)
        ws.merge_cells(start_row=3, start_column=9, end_row=4, end_column=9)
        ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
        ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=13)
        ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)
        ws.merge_cells(start_row=3, start_column=12, end_row=3, end_column=13)
        ws.merge_cells(start_row=1, start_column=14, end_row=4, end_column=14)
        ws.merge_cells(start_row=1, start_column=15, end_row=4, end_column=15)

        
        row_idx = 5
        for counter, item in enumerate(items, start=1):
            latest_relation = item.relationitemsreports_set.all().latest('last_scan_datetime') if item.relationitemsreports_set.exists() else None
            assessed_amount = latest_relation.assessed_amount if latest_relation else None
            status = latest_relation.status if latest_relation else None

            row_data = [
                counter,
                item.name,
                item.item_number,
                item.value,
                item.amount,
                item.amount * item.value,
                status.name if status is not None else "Без статуса",
                assessed_amount,
                item.assessed_value,
                assessed_amount - item.amount,
                (assessed_amount - item.amount) * item.assessed_value,
                item.amount - assessed_amount,
                (item.amount - assessed_amount) * item.assessed_value,
                item.financially_responsible_person.first_name + item.financially_responsible_person.last_name if item.financially_responsible_person else None,
                latest_relation.note if latest_relation else None
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        
        file_path = "./items_management/static/reports/items_report.xlsx"
        wb.save(file_path)

        print("Excel файл успешно сохранен:", file_path)

    except Exception as e:
        print("Ошибка при создании отчета:", e)